import random, copy
from Classes import *
from math import ceil, log2
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import math

excel_data = {}

professor_courses = {
    "John": ["English", "Art"],
    "Sarah": ["Math", "Art"],
    "Susan": ["English", "Math"],
    "Tom": ["Science", "Art"],
    "Frank": ["English", "Science"],
    "Herman": ["Math", "Art"],
    "Zed": ["English", "Math"],
    "Clint": ["Science", "Art", "Math"],
    "Peter": ["English", "Art", "Math", "Science"],
    "Bill": ["Math"],
    "Anatole": ["English"],
    "Porche": ["Math", "Science", "Art"],
    "Tandis": ["English", "Art"],
    "Sonia": ["English"],
    "Sharon": ["English", "Math"],
    "Pi": ["Math"],
    "Xi": ["Math"],
    "Chu": ["Art"],
    "Kate": ["English", "Art", "Math", "Science"],
    "Sally": ["Science"]
}

professor_availability = {
    "John": ["14:00", "15:00", "16:00", "9:30"],
    "Sarah": ["14:00", "15:00", "16:00", "10:30", "11:00"],
    "Susan": ["17:00", "14:00", "16:00", "12:00"],
    "Tom": ["18:00", "15:00", "14:00", "13:00"],
    "Frank": ["15:00", "14:00", "17:00", "14:30", "9:30"],
    "Herman": ["16:00", "18:00", "15:00", "10:00"],
    "Zed": ["15:00", "14:00", "15:30", "11:30"],
    "Clint": ["14:00", "16:00", "17:00", "16:00"],
    "Peter": ["18:00", "14:00", "16:00", "9:30", "12:30"],
    "Bill": ["15:00", "14:00", "16:00", "10:00", "13:30"],
    "Anatole": ["16:00", "15:00", "18:00", "11:00"],
    "Porche": ["17:00", "14:00", "12:30", "10:00"],
    "Tandis": ["14:00", "15:00", "9:30", "13:00"],
    "Sonia": ["17:00", "14:00", "15:00", "16:00", "10:00"],
    "Sharon": ["16:00", "14:00", "17:00", "11:30"],
    "Pi": ["15:00", "16:00", "14:00", "13:30", "10:00"],
    "Xi": ["17:00", "16:00", "15:00", "9:30", "11:00"],
    "Chu": ["15:00", "16:00", "17:00", "14:00", "13:00"],
    "Kate": ["14:00", "15:00", "16:00", "12:00"],
    "Sally": ["14:00", "15:00", "16:00", "17:00", "9:30"]
}

Group.groups = [
    Group("a", 12),
    Group("b", 12),
    Group("c", 16),
    Group("d", 16),
    Group("e", 18),
    Group("f", 18),
]

Professor.professors = [
    Professor("John"), Professor("Sarah"), Professor("Susan"), Professor("Tom"),
    Professor("Frank"), Professor("Herman"), Professor("Zed"), Professor("Clint"),
    Professor("Peter"), Professor("Bill"), Professor("Anatole"), Professor("Porche"),
    Professor("Tandis"), Professor("Sonia"), Professor("Sharon"), Professor("Pi"),
    Professor("Xi"), Professor("Chu"), Professor("Kate"), Professor("Sally"),
    Professor("John1"), Professor("Sarah1"), Professor("Susan1"), Professor("Tom1"),
    Professor("Frank1"), Professor("Herman1"), Professor("Zed1"), Professor("Clint1"),
    Professor("Peter1"), Professor("Bill1"), Professor("Anatole1"), Professor("Porche1"),
    Professor("Tandis1"), Professor("Sonia1"), Professor("Sharon1"), Professor("Pi1"),
    Professor("Xi1"), Professor("Chu1"), Professor("Kate1"), Professor("Sally1")
]

CourseClass.classes = [
    CourseClass("English"),
    CourseClass("Math"),
    CourseClass("Art"),
    CourseClass("Science"),
]

Room.rooms = [Room("A", 12), Room("B", 12), Room("C", 16), Room("D", 16), Room("E", 18), Room("F", 18)]

# M-F 14:00-20:00
# SAT,SUN: 9:30-19:00
Slot.slots = [
    Slot("14:00", "15:45", "Mon", course_class=CourseClass("Math")),
    Slot("16:00", "17:45", "Mon", course_class=CourseClass("Art")),
    Slot("18:00", "19:45", "Mon", course_class=CourseClass("Math")),
    Slot("14:00", "15:45", "Tues", course_class=CourseClass("English")),
    Slot("16:00", "17:45", "Tues", course_class=CourseClass("Science")),
    Slot("18:00", "19:45", "Tues", course_class=CourseClass("Math")),
    Slot("14:00", "15:45", "Wed", course_class=CourseClass("English")),
    Slot("16:00", "17:45", "Wed", course_class=CourseClass("Science")),
    Slot("18:00", "19:45", "Wed", course_class=CourseClass("Art")),
    Slot("14:00", "15:45", "Thu", course_class=CourseClass("English")),
    Slot("16:00", "17:45", "Thu", course_class=CourseClass("Math")),
    Slot("18:00", "19:45", "Thu", course_class=CourseClass("English")),
    Slot("14:00", "15:45", "Fri", course_class=CourseClass("Science")),
    Slot("16:00", "17:45", "Fri", course_class=CourseClass("Art")),
    Slot("18:00", "19:45", "Fri", course_class=CourseClass("Math")),
    Slot("9:30", "11:15", "Sat", course_class=CourseClass("English")),
    Slot("11:30", "13:15", "Sat", course_class=CourseClass("Science")),
    Slot("13:30", "15:15", "Sat", course_class=CourseClass("Art")),
    Slot("15:30", "17:15", "Sat", course_class=CourseClass("English")),
    Slot("17:30", "19:00", "Sat", course_class=CourseClass("Art")),
    Slot("9:30", "11:15", "Sun", course_class=CourseClass("Math")),
    Slot("11:30", "13:15", "Sun", course_class=CourseClass("Math")),
    Slot("13:30", "15:15", "Sun", course_class=CourseClass("English")),
    Slot("15:30", "17:15", "Sun", course_class=CourseClass("Science")),
    Slot("17:30", "19:00", "Sun", course_class=CourseClass("Math"))
]

# TODO
# 0.  Running Simplified Class Scheduling - Done
# 0.5 Problem Instance to Binary String - Done
# 1.  Multiple days - Done
# 2.  Class Size - Done
# 2.25 Check Selection Function - Done
# 2.5 One group can attend only one class at a time - Done
# 3.  Multiple classes - Done
# 4.  Lab - Done

# Below chromosome parts are just to teach basic

# cpg = ["000000", "010001", "100100", "111010"] # course, professor, student group pair
# lts = ["00", "01"] # lecture theatres
# slots = ["00", "01"] # time slots

# ######### Chromosome ##############
# <CourseClass, Prof, Group, Slot, LT>   #
# ###################################


max_score = None

cpg = []
lts = []
slots = []
bits_needed_backup_store = {}  # to improve performance


def bits_needed(x):
    global bits_needed_backup_store
    r = bits_needed_backup_store.get(id(x))
    if r is None:
        r = int(ceil(log2(len(x))))
        bits_needed_backup_store[id(x)] = r
    return max(r, 1)


def join_cpg_pair(_cpg):
    res = []
    for i in range(0, len(_cpg), 3):
        res.append(_cpg[i] + _cpg[i + 1] + _cpg[i + 2])
    return res


# def convert_input_to_bin():
#     global cpg, lts, slots, max_score
#
#     cpg = [CourseClass.find("hu100a"), Professor.find("mutaqi"), Group.find("a"),
#            CourseClass.find("hu100b"), Professor.find("mutaqi"), Group.find("a"),
#            CourseClass.find("mt111"), Professor.find("khalid"), Group.find("a"),
#            CourseClass.find("cs152"), Professor.find("basit"), Group.find("a"),
#            CourseClass.find("hu160"), Professor.find("mutaqi"), Group.find("b"),
#            CourseClass.find("ch110"), Professor.find("zafar"), Group.find("e"),
#            CourseClass.find("cs101"), Professor.find("basit"), Group.find("e"),
#            CourseClass.find("cs101 lab"), Professor.find("basit"), Group.find("e")
#            ]
#
#     for _c in range(len(cpg)):
#         if _c % 3:  # CourseClass
#             cpg[_c] = (bin(cpg[_c])[2:]).rjust(bits_needed(CourseClass.classes), '0')
#         elif _c % 3 == 1:  # Professor
#             cpg[_c] = (bin(cpg[_c])[2:]).rjust(bits_needed(Professor.professors), '0')
#         else:  # Group
#             cpg[_c] = (bin(cpg[_c])[2:]).rjust(bits_needed(Group.groups), '0')
#
#     cpg = join_cpg_pair(cpg)
#     for r in range(len(Room.rooms)):
#         lts.append((bin(r)[2:]).rjust(bits_needed(Room.rooms), '0'))
#
#     for t in range(len(Slot.slots)):
#         slots.append((bin(t)[2:]).rjust(bits_needed(Slot.slots), '0'))
#
#     # print(cpg)
#     max_score = (len(cpg) - 1) * 3 + len(cpg) * 3

def convert_input_to_bin():
    global cpg, lts, slots, max_score

    cpg = []
    # Room.assigned_slots = {}
    for professor in Professor.professors:
        for course in professor_courses.get(professor.name, []):
            group = random.choice(Group.groups)

            # Consider professor availability
            available_slots = [Slot.slots[i] for i, slot in enumerate(Slot.slots)
                               if slot.start in professor_availability.get(professor.name, [])]
            if not available_slots: continue

            assigned_room = False
            num_tries = 500
            # print("Before assign", Room.assigned_slots)
            for i in range(num_tries):
                slot = random.choice(available_slots)
                room = random.choice(Room.rooms)
                if room.name not in Room.assigned_slots:
                    Room.assigned_slots[room.name] = []
                if slot.get_key() not in Room.assigned_slots[room.name]:
                    Room.assigned_slots[room.name].append(slot.get_key())
                    assigned_room = True
                    break
            if not assigned_room:
                continue
            # print("after assign", Room.assigned_slots)

            cpg.extend([
                (bin(CourseClass.find(course))[2:]).rjust(bits_needed(CourseClass.classes), '0'),
                (bin(Professor.find(professor.name))[2:]).rjust(bits_needed(Professor.professors), '0'),
                (bin(Group.find(group.name))[2:]).rjust(bits_needed(Group.groups), '0')
            ])
            lts.append((bin(Room.find(room.name))[2:]).rjust(bits_needed(Room.rooms), '0'))
            slots.append((bin(Slot.slots.index(slot))[2:]).rjust(bits_needed(Slot.slots), '0'))

    cpg = join_cpg_pair(cpg)
    max_score = (len(cpg) - 1) * 3 + len(cpg) * 3


def course_bits(chromosome):
    i = 0

    return chromosome[i:i + bits_needed(CourseClass.classes)]


def professor_bits(chromosome):
    i = bits_needed(CourseClass.classes)

    return chromosome[i: i + bits_needed(Professor.professors)]


def group_bits(chromosome):
    i = bits_needed(CourseClass.classes) + bits_needed(Professor.professors)

    return chromosome[i:i + bits_needed(Group.groups)]


def slot_bits(chromosome):
    i = bits_needed(CourseClass.classes) + bits_needed(Professor.professors) + \
        bits_needed(Group.groups)

    return chromosome[i:i + bits_needed(Slot.slots)]


def lt_bits(chromosome):
    i = bits_needed(CourseClass.classes) + bits_needed(Professor.professors) + \
        bits_needed(Group.groups) + bits_needed(Slot.slots)

    return chromosome[i: i + bits_needed(Room.rooms)]


def slot_clash(a, b):
    if slot_bits(a) == slot_bits(b):
        return 1
    return 0


# checks that a faculty member teaches only one course at a time.
def faculty_member_one_class(chromosome):
    scores = 0
    for i in range(len(chromosome) - 1):  # select one cpg pair
        clash = False
        for j in range(i + 1, len(chromosome)):  # check it with all other cpg pairs
            if slot_clash(chromosome[i], chromosome[j]) \
                    and professor_bits(chromosome[i]) == professor_bits(chromosome[j]):
                clash = True
                # print("These prof. have clashes")
                # print_chromosome(chromosome[i])
                # print_chromosome(chromosome[j])
        if not clash:
            scores = scores + 1
    return scores


# check that a group member takes only one class at a time.
def group_member_one_class(chromosomes):
    scores = 0

    for i in range(len(chromosomes) - 1):
        clash = False
        for j in range(i + 1, len(chromosomes)):
            if slot_clash(chromosomes[i], chromosomes[j]) and \
                    group_bits(chromosomes[i]) == group_bits(chromosomes[j]):
                # print("These classes have slot/lts clash")
                # print_chromosome(chromosomes[i])
                # print_chromosome(chromosomes[j])
                # print("____________")
                clash = True
                break
        if not clash:
            # print("These classes have no slot/lts clash")
            # print_chromosome(chromosomes[i])
            # print_chromosome(chromosomes[j])
            # print("____________")
            scores = scores + 1
    return scores


# checks that a course is assigned to an available classroom.
def use_spare_classroom(chromosome):
    scores = 0
    for i in range(len(chromosome) - 1):  # select one cpg pair
        clash = False
        for j in range(i + 1, len(chromosome)):  # check it with all other cpg pairs
            if slot_clash(chromosome[i], chromosome[j]) and lt_bits(chromosome[i]) == lt_bits(chromosome[j]):
                # print("These classes have slot/lts clash")
                # printChromosome(chromosome[i])
                # printChromosome(chromosome[j])
                clash = True
        if not clash:
            scores = scores + 1
    return scores


# checks that the classroom capacity is large enough for the classes that
# are assigned to that classroom.
def classroom_size(chromosomes):
    scores = 0
    for _c in chromosomes:
        if Group.groups[int(group_bits(_c), 2)].size <= Room.rooms[int(lt_bits(_c), 2)].size:
            scores = scores + 1
    return scores


# check that room is appropriate for particular class/lab
def appropriate_room(chromosomes):
    scores = 0
    for _c in chromosomes:
        if CourseClass.classes[int(course_bits(_c), 2)].is_lab == Room.rooms[int(lt_bits(_c), 2)].is_lab:
            scores = scores + 1
    return scores


# check that lab is allocated appropriate time slot
def appropriate_timeslot(chromosomes):
    scores = 0
    for _c in chromosomes:
        if CourseClass.classes[int(course_bits(_c), 2)].is_lab == Slot.slots[int(slot_bits(_c), 2)].is_lab_slot:
            scores = scores + 1
    return scores


def evaluate(chromosomes):
    global max_score
    score = 0
    score += faculty_availability(chromosomes)
    score = score + use_spare_classroom(chromosomes)
    score = score + faculty_member_one_class(chromosomes)
    score = score + classroom_size(chromosomes)
    score = score + group_member_one_class(chromosomes)
    score = score + appropriate_room(chromosomes)
    score = score + appropriate_timeslot(chromosomes)
    return score / max_score

def faculty_availability(chromosomes):
    scores = 0
    for _c in chromosomes:
        professor_name = Professor.professors[int(professor_bits(_c), 2)].name
        slot_time = Slot.slots[int(slot_bits(_c), 2)].start
        if slot_time in professor_availability.get(professor_name, []):
            scores += 1
    return scores

def cost(solution):
    # solution would be an array inside an array
    # it is because we use it as it is in genetic algorithms
    # too. Because, GA require multiple solutions i.e population
    # to work.
    return 1 / float(evaluate(solution))


def init_population(n):
    global cpg, lts, slots
    chromosomes = []
    for _n in range(n):
        chromosome = []
        for _c in cpg:
            chromosome.append(_c + random.choice(slots) + random.choice(lts))
        chromosomes.append(chromosome)
    return chromosomes


# Modified Combination of Row_reselect, Column_reselect
def mutate(chromosome):
    rand_slot = random.choice(slots)
    rand_lt = random.choice(lts)

    a = random.randint(0, len(chromosome) - 1)

    # Respect professor availability in mutation
    professor_name = Professor.professors[int(professor_bits(chromosome[a]), 2)].name
    available_slots = [Slot.slots[i] for i, slot in enumerate(Slot.slots)
                       if slot.start in professor_availability.get(professor_name, [])]
    if not available_slots: return
    rand_slot = random.choice(available_slots).start

    chromosome[a] = course_bits(chromosome[a]) + professor_bits(chromosome[a]) + \
                    group_bits(chromosome[a]) + rand_slot + lt_bits(chromosome[a])

    # print("After mutation: ", end="")
    # printChromosome(chromosome)


def crossover(population):
    a = random.randint(0, len(population) - 1)
    b = random.randint(0, len(population) - 1)
    cut = random.randint(0, len(population[0]))  # assume all chromosome are of same len
    population.append(population[a][:cut] + population[b][cut:])


def selection(population, n):
    population.sort(key=evaluate, reverse=True)
    while len(population) > n:
        population.pop()


def print_chromosome(chromosome):
    # data = {
    #     "[room]": [
    #         ["Time", "Mon", "Tues", "Wed", "Thu", "Fri", "Sat", "Sun"]
    #     ]
    # }
    day_to_index = {
        "Mon": 1,
        "Tues": 2,
        "Wed": 3,
        "Thu": 4,
        "Fri": 5,
        "Sat": 6,
        "Sun": 7
    }
    course = CourseClass.classes[int(course_bits(chromosome), 2)]
    professor = Professor.professors[int(professor_bits(chromosome), 2)]
    group = Group.groups[int(group_bits(chromosome), 2)]
    slot = Slot.slots[int(slot_bits(chromosome), 2)]
    room = Room.rooms[int(lt_bits(chromosome), 2)]

    if room.name not in excel_data:
        excel_data[room.name] = [
            ["Time", "Mon", "Tues", "Wed", "Thu", "Fri", "Sat", "Sun"]
        ]

    row = [''] * 8
    row[0] = "{start}-{end}".format(start=slot.start, end=slot.end)
    row[day_to_index[slot.day]] = "{subject}: {professor}".format(subject=course.code, professor=professor.name)
    excel_data[room.name].append(row)
    print(CourseClass.classes[int(course_bits(chromosome), 2)], " | ",
          Professor.professors[int(professor_bits(chromosome), 2)], " | ",
          Group.groups[int(group_bits(chromosome), 2)], " | ",
          Slot.slots[int(slot_bits(chromosome), 2)], " | ",
          Room.rooms[int(lt_bits(chromosome), 2)])


def write_excel():
    """
    {'E': [
            ['Time', 'Mon', 'Tues', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'],
            ['9:30-11:15', '', '', '', '', '', 'English: John', '']
        ]
    }
    openpyxl_width = excel_width * 1.143

    13.55598
    """
    blue_fill = PatternFill(start_color="365f92", end_color="365f92", fill_type="solid")
    light_blue_fill = PatternFill(start_color="dce6f1", end_color="dce6f1", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    wb = Workbook()
    # ws = wb.active
    for key, value in sorted(excel_data.items()):
        wb.create_sheet(key)
        ws = wb[key]
        is_first_row = True
        excel_row_index = 1
        for row in excel_data[key]:
            ws.append(row)
            if is_first_row:
                is_first_row = False
                for col in range(1, len(row) + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.fill = blue_fill
                    cell.font = white_font
            else:
                for col in range(2, len(row) + 1):
                    cell = ws.cell(row=excel_row_index, column=col)
                    if len(str(cell.value)) > 0:
                        cell.fill = light_blue_fill
            excel_row_index += 1

        # Get all values from column A
        values = [cell.value for cell in ws['A'] if cell.value is not None]
        # Sort the values in descending order
        values.sort(reverse=True)
        # Update the worksheet with the sorted values
        for idx, value in enumerate(values, start=1):
            ws[f'A{idx}'].value = value

        adjust_column_width(ws)
        ws.column_dimensions['A'].width = 13.55598

    default_sheet = wb['Sheet']
    wb.remove(default_sheet)
    wb.save("Class_Schedule_colored.xlsx")

def adjust_column_width(worksheet):
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)  # Adding a little extra space
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

# Simple Searching Neighborhood
# It randomly changes timeslot of a class/lab
def ssn(solution):
    rand_slot = random.choice(slots)
    rand_lt = random.choice(lts)

    a = random.randint(0, len(solution) - 1)

    new_solution = copy.deepcopy(solution)
    new_solution[a] = course_bits(solution[a]) + professor_bits(solution[a]) + \
                      group_bits(solution[a]) + rand_slot + lt_bits(solution[a])
    return [new_solution]


# Swapping Neighborhoods
# It randomy selects two classes and swap their time slots
def swn(solution):
    a = random.randint(0, len(solution) - 1)
    b = random.randint(0, len(solution) - 1)
    new_solution = copy.deepcopy(solution)
    temp = slot_bits(solution[a])
    new_solution[a] = course_bits(solution[a]) + professor_bits(solution[a]) + \
                      group_bits(solution[a]) + slot_bits(solution[b]) + lt_bits(solution[a])

    new_solution[b] = course_bits(solution[b]) + professor_bits(solution[b]) + \
                      group_bits(solution[b]) + temp + lt_bits(solution[b])
    # print("Diff", solution)
    # print("Meiw", new_solution)
    return [new_solution]


def acceptance_probability(old_cost, new_cost, temperature):
    if new_cost < old_cost:
        return 1.0
    else:
        return math.exp((old_cost - new_cost) / temperature)


def simulated_annealing():
    alpha = 0.9
    T = 1.0
    T_min = 0.00001

    convert_input_to_bin()
    population = init_population(1)  # as simulated annealing is a single-state method
    old_cost = cost(population[0])
    # print("Cost of original random solution: ", old_cost)
    # print("Original population:")
    # print(population)

    for __n in range(500):
        new_solution = swn(population[0])
        new_solution = ssn(population[0])
        new_cost = cost(new_solution[0])
        ap = acceptance_probability(old_cost, new_cost, T)
        if ap > random.random():
            population = new_solution
            old_cost = new_cost
        T = T * alpha
    # print(population)
    # print("Cost of altered solution: ", cost(population[0]))
    print("\n------------- Simulated Annealing --------------\n")
    for lec in population[0]:
        print_chromosome(lec)
    print("Score: ", evaluate(population[0]))


def genetic_algorithm():
    generation = 0
    convert_input_to_bin()
    population = init_population(3)

    # print("Original population:")
    # print(population)
    print("\n------------- Genetic Algorithm --------------\n")
    while True:

        # if termination criteria are satisfied, stop.
        if evaluate(max(population, key=evaluate)) == 1 or generation == 500:
            print("Generations:", generation)
            print("Best Chromosome fitness value", evaluate(max(population, key=evaluate)))
            print("Best Chromosome: ", max(population, key=evaluate))
            for lec in max(population, key=evaluate):
                print_chromosome(lec)
            break

        # Otherwise continue
        else:
            for _c in range(len(population)):
                crossover(population)
                selection(population, 5)

                # selection(population[_c], len(cpg))
                mutate(population[_c])

        generation = generation + 1
        # print("Gen: ", generation)

    # print("Population", len(population))


def main():
    random.seed()
    # genetic_algorithm()
    simulated_annealing()
    print(excel_data)
    write_excel()


main()